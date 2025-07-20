import urllib.request
import re
import datetime
import openpyxl

def get_rankings_from_url(url, keyword):
    try:
        res = urllib.request.urlopen(url, timeout=15)
        html = res.read().decode('utf-8')
    except Exception as e:
        print(f"{keyword}ページ取得エラー: {e}")
        return ['ランキング情報なし']

    # 抽出ブロックの切り出し
    start = html.find("Amazon 売れ筋ランキング:")
    end = html.find("カスタマーレビュー", start)
    rankings = []
    if start != -1 and end != -1:
        block = html[start:end]
        # HTMLタグ除去
        block = re.sub('<.*?>', '', block)
        block = re.sub(r'\s+', ' ', block)
        rankings = [r.strip() for r in block.split('-') if r.strip()]
    else:
        rankings = ['ランキング情報なし']
    return rankings

# 日付＋ランキング
now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M')

# 1. 紙書籍版（Sheet1に保存）
normal_url = 'https://www.amazon.co.jp/dp/4798183180'
normal_rankings = [now] + get_rankings_from_url(normal_url, '紙書籍')

# 2. Kindle版（Kindleシートに保存）
kindle_url = 'https://www.amazon.co.jp/dp/B0CYPMKYM3'
kindle_rankings = [now] + get_rankings_from_url(kindle_url, 'Kindle')

# ===== Excelに書き込み =====
try:
    wb = openpyxl.load_workbook('C:/Users/Hana/python/amazonranking_matome.xlsx')

    def write_to_sheet(sheet_name, data):
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        next_row = ws.max_row + 1
        for col_index, value in enumerate(data, start=1):
            ws.cell(row=next_row, column=col_index, value=value)

    write_to_sheet('Sheet1', normal_rankings)
    write_to_sheet('Kindle', kindle_rankings)

    wb.save('C:/Users/Hana/python/amazonranking_matome.xlsx')
    print("Excelに保存しました。")

except Exception as e:
    print(f"Excel更新エラー: {e}")
