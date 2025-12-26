import urllib.request
import re
import datetime
import pytz
import openpyxl
import time
import os
import json

import gspread
from gspread.exceptions import WorksheetNotFound
from oauth2client.service_account import ServiceAccountCredentials

BASE_DIR = os.path.dirname(__file__)
SHEET_NAME = 'Amazon 売れ筋ランキング'

def log(msg):
    timestamp = time.strftime('%H:%M:%S')
    print(f"[{timestamp}] {msg}")
    try:
        with open('amazonranking_log.txt', 'a', encoding='utf-8') as f:
            f.write(f"[{timestamp}] {msg}\n")
    except Exception as e:
        print(f"ログ書き込みエラー: {e}")

def extract_rankings_from_html(html, keyword, expected_len):
    """
    Amazon商品ページのHTMLから売れ筋ランキング情報を抽出する。
    """
    rankings = []

    # 売れ筋ランキング部分を検出
    markers = ["Amazon 売れ筋ランキング:", "Amazon 売れ筋ランキング", "売れ筋ランキング:"]
    start = -1
    used_marker = None
    for m in markers:
        start = html.find(m)
        if start != -1:
            used_marker = m
            break

    if start == -1:
        log(f"{keyword}ランキング情報が見つかりません")
        return ['-'] * expected_len

    # 見出し文字列そのものを block に含めないように調整
    start = start + len(used_marker)
    end = html.find("カスタマーレビュー", start)
    if end == -1:
        end = len(html)
    block = html[start:end]

    # HTMLタグ除去＋空白正規化
    block = re.sub(r'<.*?>', '', block)
    block = re.sub(r'\s+', ' ', block).strip()

    # 不要語削除（「～の売れ筋ランキングを見る」などのリンクテキストを削除）
    block = re.sub(r'\(?.*?の売れ筋ランキングを見る\)?', '', block)
    block = re.sub(r'.*?の売れ筋ランキングを見る', '', block)

    log(f"{keyword}処理前のブロック: {block[:200]}")

    # 「カテゴリ - 順位」形式を抽出
    pattern = r'([^\-:：]{2,80}?)\s*[-−]\s*(\d{1,3}(?:,\d{3})*位)'
    matches = re.findall(pattern, block)

    if not matches:
        log(f"{keyword}ランキングパターンに一致なし")
        return ['-'] * expected_len

    for name, rank in matches:
        name = name.strip()
        rank = rank.strip()
        # ノイズ除去
        if "Amazon" in name or "見る" in name:
            continue
        text = f"{rank}{name}"
        text = re.sub(r'\(\s*\)', '', text)
        rankings.append(text)

    # 指定された列数に合わせる
    if len(rankings) < expected_len:
        rankings += ['-'] * (expected_len - len(rankings))
    else:
        rankings = rankings[:expected_len]

    log(f"{keyword}ランキング抽出完了: {rankings}")
    return rankings

def get_rankings_from_url(url, keyword, expected_len):
    """URLからHTMLを取得し、指定された件数のランキングを抽出して返す。"""
    log(f"{keyword}ページ取得開始: {url}")
    try:
        # Amazonの制限を避けるためUser-Agentを設定
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as res:
            html = res.read().decode('utf-8')
        log(f"{keyword}ページ取得完了")
    except Exception as e:
        log(f"{keyword}ページ取得エラー: {e}")
        return ['-'] * expected_len

    return extract_rankings_from_html(html, keyword, expected_len)

def save_to_excel_with_retry(excel_path, row_data, max_retries=3):
    for attempt in range(max_retries):
        try:
            log(f"Excel保存試行 {attempt + 1}/{max_retries}")
            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
            else:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            if SHEET_NAME not in wb.sheetnames:
                if 'Sheet1' in wb.sheetnames and len(wb.sheetnames) == 1:
                    wb['Sheet1'].title = SHEET_NAME
                else:
                    wb.create_sheet(SHEET_NAME)
            ws = wb[SHEET_NAME]
            ws.append(row_data)
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            wb.save(excel_path)
            log("Excel書き込み完了")
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                time.sleep(5)
            else: return False
        except Exception: return False
    return False

def append_to_google_sheet(row_data):
    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    if not creds_json: return
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    try:
        info = json.loads(creds_json)
        creds = ServiceAccountCredentials.from_json_keyfile_dict(info, scope)
        client = gspread.authorize(creds)
        
        # ご指定いただいたスプレッドシートID
        SPREADSHEET_ID = "1DSn3IK9ebd0apbqe2WIXKaRGrDVg7XhaK1jlQZrjBk8"
        
        workbook = client.open_by_key(SPREADSHEET_ID)
        try:
            worksheet = workbook.worksheet(SHEET_NAME)
        except WorksheetNotFound:
            worksheet = workbook.add_worksheet(title=SHEET_NAME, rows="10000", cols="20")
        worksheet.append_row(row_data, value_input_option='USER_ENTERED')
        worksheet.sort((1, 'des'))
        log("Googleスプレッドシートに追記・ソート完了")
    except Exception as e:
        log(f"Googleスプレッドシート書き込みエラー: {e}")

# -------------------------------
# メイン処理
# -------------------------------
log("処理開始")
JST = pytz.timezone('Asia/Tokyo')
now = datetime.datetime.now(JST).replace(minute=0, second=0, microsecond=0).strftime('%Y/%m/%d %H:%M')

# URL設定（元のパターンの書き方）
normal_url = 'https://www.amazon.co.jp/gp/product/4798183180/'
kindle_url = 'https://www.amazon.co.jp/gp/product/B0CYPMKYM3/'
audible_url = 'https://www.amazon.co.jp/gp/product/B0G66DNXDH/'

# 各カテゴリの取得件数を指定
# 紙書籍: 4件, Kindle: 2件, Audible: 4件
normal_rankings = get_rankings_from_url(normal_url, '紙書籍', expected_len=4)
kindle_rankings = get_rankings_from_url(kindle_url, 'Kindle', expected_len=2)
audible_rankings = get_rankings_from_url(audible_url, 'Audible', expected_len=4)

row_data = [now] + normal_rankings + kindle_rankings + audible_rankings
log(f"構築された行データ: {row_data}")

# 保存処理
append_to_google_sheet(row_data)
excel_path = os.path.join(BASE_DIR, 'amazonranking_matome.xlsx')
save_to_excel_with_retry(excel_path, row_data)

log("処理完了")